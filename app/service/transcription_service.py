# coding:utf-8
"""听写服务 - 核心转录功能"""
import os
import re
import sys
import subprocess
from pathlib import Path
from typing import Optional, Dict, Any, List
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment

from ..common.config import cfg
from ..common.model_scanner import modelScanner


def get_ffmpeg_path() -> str:
    """
    获取 ffmpeg 可执行文件路径
    
    Returns:
        ffmpeg 的完整路径
    """
    # 获取当前文件所在目录
    current_dir = Path(__file__).parent.parent  # app 目录
    
    # 尝试多个可能的路径
    possible_paths = [
        current_dir / 'tools' / 'ffmpeg.exe',           # app/tools/ffmpeg.exe
        current_dir / 'common' / 'tools' / 'ffmpeg.exe',  # app/common/tools/ffmpeg.exe
        Path('ffmpeg.exe'),                             # 当前目录
    ]
    
    for ffmpeg_path in possible_paths:
        if ffmpeg_path.exists():
            print(f"[INFO] 找到 ffmpeg: {ffmpeg_path}")
            return str(ffmpeg_path)
    
    # 否则尝试使用系统 PATH 中的 ffmpeg
    print("[INFO] 使用系统 PATH 中的 ffmpeg")
    return 'ffmpeg'


class WhisperEngine:
    """Whisper 引擎类型"""
    GGML = "ggml"  # whisper.cpp
    FASTER_WHISPER = "faster-whisper"
    NONE = "不进行听写"


class OutputFormat:
    """输出格式"""
    SRT_ORIGINAL = "原文SRT"
    SRT_BILINGUAL = "双语SRT"
    LRC_ORIGINAL = "原文LRC"
    TXT_ORIGINAL = "原文TXT"
    TXT_BILINGUAL = "双语TXT"
    XLSX_ORIGINAL = "原文XLSX"
    XLSX_BILINGUAL = "双语XLSX"


class TranscriptionService:
    """听写服务 - 核心转录功能"""
    
    def __init__(self):
        self._available_models = []  # 可用模型列表
        self._param_template = ""    # 参数模板
        self._available = False      # 服务是否可用
        self._check_availability()
        self._scan_models()
    
    def _scan_models(self):
        """扫描可用的 Whisper 模型"""
        print("\n[模型扫描] 开始扫描 Faster-Whisper 模型...")
        
        # 扫描 Faster-Whisper 模型
        faster_models = modelScanner.scan_faster_whisper_models()
        self._available_models = faster_models
        
        # 读取参数模板
        self._param_template = modelScanner.read_param_template()
        
        print(f"[模型扫描] 扫描完成，共发现 {len(self._available_models)} 个模型\n")
    
    def get_available_models(self) -> List[str]:
        """
        获取可用的模型列表
        
        Returns:
            模型名称列表
        """
        return self._available_models.copy()
    
    def _check_availability(self):
        """检查服务依赖是否可用"""
        # 检查 ffmpeg 是否可用
        try:
            ffmpeg_path = get_ffmpeg_path()
            subprocess.run([ffmpeg_path, '-version'], 
                         capture_output=True, 
                         check=True,
                         creationflags=0x08000000 if sys.platform == 'win32' else 0)
            self._available = True
            print(f"[INFO] 听写服务已就绪 (ffmpeg: {ffmpeg_path})")
        except (subprocess.CalledProcessError, FileNotFoundError) as e:
            self._available = False
            print(f"[WARNING] ffmpeg 未找到，听写服务不可用 (错误: {e})")
    
    def isAvailable(self) -> bool:
        """检查服务是否可用"""
        return self._available
    
    def transcribe(self, input_path: str, **kwargs) -> Optional[Dict[str, Any]]:
        """
        执行听写转录
        
        Args:
            input_path: 输入文件路径（视频/音频文件）
            **kwargs: 额外参数
                - whisper_model: Whisper 模型文件名
                - language: 源语言（ja, en, zh 等）
                - output_format: 输出格式
                    * "原文SRT" - SRT 字幕文件
                    * "双语SRT" - 双语 SRT（需配合翻译）
                    * "原文LRC" - LRC 歌词文件
                    * "原文TXT" - 纯文本文件
                    * "双语TXT" - 双语文本（需配合翻译）
                    * "原文XLSX" - Excel 表格
                    * "双语XLSX" - 双语 Excel（需配合翻译）
                - whisper_params: Whisper 参数
                - faster_whisper_params: Faster-Whisper 参数
                - translated_srt: 翻译后的 SRT 文件路径（用于双语格式）
                - include_timestamp: 是否包含时间戳（默认 True）
                - split_parts: 均分人数（默认 0，表示不均分）
                - save_folder: 保存文件夹路径（默认为输入文件所在目录）
        
        Returns:
            结果字典，包含 output_path 和 srt_path，如果失败返回 None
        """
        if not self.isAvailable():
            print("[ERROR] 听写服务不可用")
            return None
        
        input_file = Path(input_path)
        if not input_file.exists():
            print(f"[ERROR] 输入文件不存在: {input_path}")
            return None
        
        # 获取参数
        whisper_model = kwargs.get('whisper_model', WhisperEngine.NONE)
        language = kwargs.get('language', 'ja')
        output_format = kwargs.get('output_format', OutputFormat.SRT_ORIGINAL)
        
        # 输出任务开始日志
        print("\n" + "#"*80)
        print(f"# 听写任务开始")
        print(f"# 文件名: {input_file.name}")
        print(f"# 开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("#"*80 + "\n")
        sys.stdout.flush()
        
        # 输出配置信息
        print(f"[任务配置]")
        print(f"  - 输入文件: {input_file}")
        print(f"  - Whisper 模型: {whisper_model}")
        print(f"  - 源语言: {language}")
        print(f"  - 输出格式: {output_format}")
        print()
        sys.stdout.flush()
        
        try:
            # 如果是 SRT 文件，直接转换格式
            if input_file.suffix.lower() == '.srt':
                print("[INFO] 检测到 SRT 文件，进行格式转换...")
                return self._convert_srt_format(input_file, output_format, kwargs)
            
            # 检查是否跳过听写
            if whisper_model == WhisperEngine.NONE:
                print("[WARNING] 未选择 Whisper 模型，跳过听写")
                return None
            
            # 1. 提取音频
            print("[INFO] 正在提取音频...")
            wav_file = self._extract_audio(input_file)
            
            # 2. 执行 Whisper 听写
            print(f"[INFO] 正在进行语音识别...（模型: {whisper_model}）")
            srt_file = self._run_whisper(
                wav_file=wav_file,
                whisper_model=whisper_model,
                language=language,
                whisper_params=kwargs
            )
            
            # 3. 生成输出文件
            print("[INFO] 正在生成输出文件...")
            output_path = self._generate_output(
                srt_file=srt_file,
                input_file=input_file,
                output_format=output_format,
                kwargs=kwargs
            )
            
            # 4. 清理临时文件
            if wav_file.exists():
                wav_file.unlink()
                print(f"[清理] 删除临时文件: {wav_file.name}")
                sys.stdout.flush()
            
            # 输出成功日志
            print("\n" + "#"*80)
            print(f"# 听写任务完成 ✓")
            print(f"# 文件名: {input_file.name}")
            print(f"# 输出路径: {output_path}")
            print(f"# 完成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print("#"*80 + "\n")
            sys.stdout.flush()
            
            return {
                'output_path': str(output_path),
                'srt_path': str(srt_file)
            }
            
        except MemoryError as e:
            # 专门处理内存错误
            print("\n" + "#"*80)
            print(f"# 听写任务失败 ✗ - 内存不足")
            print(f"# 文件名: {input_file.name}")
            print(f"# 错误类型: MemoryError (内存分配失败)")
            print(f"# 建议：")
            print(f"#   1. 使用更小的模型（如 small 或 base）")
            print(f"#   2. 关闭其他占用内存的程序")
            print(f"#   3. 处理较短的音频片段")
            print(f"#   4. 升级系统内存")
            print(f"# 失败时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print("#"*80 + "\n")
            sys.stdout.flush()
            raise RuntimeError("内存不足，无法完成转录。请尝试使用更小的模型或处理较短的音频文件。")
        
        except FileNotFoundError as e:
            # 处理文件不存在错误
            print("\n" + "#"*80)
            print(f"# 听写任务失败 ✗ - 文件未找到")
            print(f"# 文件名: {input_file.name}")
            print(f"# 错误信息: {str(e)}")
            print(f"# 失败时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print("#"*80 + "\n")
            sys.stdout.flush()
            raise
        
        except Exception as e:
            # 处理其他错误
            error_msg = str(e)
            print("\n" + "#"*80)
            print(f"# 听写任务失败 ✗")
            print(f"# 文件名: {input_file.name}")
            print(f"# 错误类型: {type(e).__name__}")
            print(f"# 错误信息: {error_msg}")
            
            # 检查是否是 whisper-faster 的 MemoryError
            if "MemoryError" in error_msg or "bad allocation" in error_msg:
                print(f"# ")
                print(f"# 这是内存分配错误，建议：")
                print(f"#   - 使用更小的模型")
                print(f"#   - 已自动使用 int8 精度以降低内存占用")
                print(f"#   - 如果仍然失败，请尝试 base 或 small 模型")
            
            print(f"# 失败时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print("#"*80 + "\n")
            sys.stdout.flush()
            raise
    
    def _extract_audio(self, input_file: Path) -> Path:
        """
        提取音频并转换为 16k 采样率的 WAV 文件
        
        Args:
            input_file: 输入文件路径
            
        Returns:
            提取的 WAV 文件路径
        """
        wav_file = self._get_unique_filename(input_file.with_suffix('.wav'))
        
        ffmpeg_path = get_ffmpeg_path()
        cmd = [
            ffmpeg_path, '-y',
            '-i', str(input_file),
            '-acodec', 'pcm_s16le',
            '-ac', '1',
            '-ar', '16000',
            str(wav_file)
        ]
        
        # 输出分隔线和标题（会自动写入 log.txt）
        print("\n" + "="*80)
        print(f"[FFmpeg] 开始提取音频: {input_file.name}")
        print(f"[FFmpeg] 命令: {' '.join(cmd)}")
        print("="*80 + "\n")
        sys.stdout.flush()  # 立即刷新到文件
        
        creationflags = 0x08000000 if sys.platform == 'win32' else 0
        
        # 设置子进程环境变量，强制使用 UTF-8
        env = os.environ.copy()
        env['PYTHONIOENCODING'] = 'utf-8'
        env['PYTHONUTF8'] = '1'
        
                
        # 关键改动：让子进程直接继承父进程的 stdout/stderr
        # 不使用 PIPE，FFmpeg 的输出会直接写入到已重定向的 sys.stdout（即 log.txt）
        process = subprocess.Popen(
            cmd,
            stdout=sys.stdout,      # 直接继承父进程的 stdout（已重定向到 log.txt）
            stderr=sys.stdout,      # stderr 也输出到 stdout
            creationflags=creationflags,
            env=env                 # 传递 UTF-8 环境变量
        )
        
        # 等待进程完成
        return_code = process.wait()
        
        # 输出结果（会自动写入 log.txt）
        print("\n" + "="*80)
        if return_code != 0 or not wav_file.exists():
            print(f"[FFmpeg] 音频提取失败，返回码: {return_code}")
            print("="*80 + "\n")
            sys.stdout.flush()
            raise RuntimeError(f"音频提取失败，返回码: {return_code}")
        
        print(f"[FFmpeg] 音频提取完成: {wav_file.name}")
        print("="*80 + "\n")
        sys.stdout.flush()
        
        return wav_file
    
    def _run_whisper(self, wav_file: Path, whisper_model: str, 
                     language: str, whisper_params: Dict[str, Any]) -> Path:
        """
        运行 Whisper 听写
        
        Args:
            wav_file: WAV 音频文件路径
            whisper_model: Whisper 模型名称
            language: 源语言
            whisper_params: Whisper 参数字典
            
        Returns:
            生成的 SRT 文件路径
        """
        output_base = wav_file.with_suffix('')
        srt_file = output_base.with_suffix('.srt')
        
        if whisper_model.startswith('ggml'):
            # 使用 whisper.cpp
            cmd = self._prepare_whisper_cpp_command(
                whisper_model, wav_file, language, whisper_params
            )
        elif whisper_model.startswith('faster-whisper'):
            # 使用 faster-whisper
            cmd = self._prepare_faster_whisper_command(
                whisper_model, wav_file, language, whisper_params
            )
        else:
            raise ValueError(f"不支持的 Whisper 模型类型: {whisper_model}")
        
        # 输出分隔线和标题（会自动写入 log.txt）
        print("\n" + "="*80)
        print(f"[Whisper] 开始语音识别")
        print(f"[Whisper] 模型: {whisper_model}")
        print(f"[Whisper] 语言: {language}")
        print(f"[Whisper] 输入文件: {wav_file.name}")
        print("="*80 + "\n")
        sys.stdout.flush()
        
        # 执行命令
        creationflags = 0x08000000 if sys.platform == 'win32' else 0
        
        # 设置子进程环境变量，强制使用 UTF-8
        env = os.environ.copy()
        env['PYTHONIOENCODING'] = 'utf-8'
        env['PYTHONUTF8'] = '1'
        
        # 关键改动：让子进程直接继承父进程的 stdout/stderr
        process = subprocess.Popen(
            cmd,
            stdout=sys.stdout,      # 直接继承父进程的 stdout（已重定向到 log.txt）
            stderr=sys.stdout,      # stderr 也输出到 stdout
            creationflags=creationflags,
            env=env                 # 传递 UTF-8 环境变量
        )
        
        # 等待进程完成
        return_code = process.wait()
        
        # 输出结果（会自动写入 log.txt）
        print("\n" + "="*80)
        if return_code != 0:
            print(f"[Whisper] 听写失败，返回码: {return_code}")
            print("="*80 + "\n")
            sys.stdout.flush()
            raise RuntimeError(f"Whisper 听写失败，返回码: {return_code}")
        
        if not srt_file.exists():
            print(f"[Whisper] 错误: 未生成 SRT 文件: {srt_file}")
            print("="*80 + "\n")
            sys.stdout.flush()
            raise RuntimeError(f"未生成 SRT 文件: {srt_file}")
        
        print(f"[Whisper] 语音识别完成: {srt_file.name}")
        print("="*80 + "\n")
        sys.stdout.flush()
        
        return srt_file
    
    def _prepare_whisper_cpp_command(self, model: str, wav_file: Path, 
                                     language: str, whisper_params: Dict[str, Any]) -> list:
        """准备 whisper.cpp 命令"""
        params = whisper_params.get('whisper_params', '')
        
        # 基础命令
        cmd = [
            'whisper/main.exe' if sys.platform == 'win32' else 'whisper/main',
            '-m', f'whisper/{model}',
            '-l', language,
            '-f', str(wav_file.with_suffix('')),
            '-osrt'
        ]
        
        # 添加额外参数
        if params:
            cmd.extend(params.split())
        
        return cmd
    
    def _prepare_faster_whisper_command(self, model: str, wav_file: Path,
                                       language: str, whisper_params: Dict[str, Any]) -> list:
        """准备 faster-whisper 命令（使用可执行文件）"""
        # 检查模型名称是否以 faster-whisper 开头
        if model.startswith('faster-whisper-'):
            # 去掉前缀 "faster-whisper-"（15个字符），得到实际模型名称
            actual_model_name = model[15:]
        elif model.startswith('faster-whisper'):
            # 兼容处理：去掉 "faster-whisper" 后的内容
            actual_model_name = model.replace('faster-whisper', '').lstrip('-')
        else:
            # 直接使用模型名称
            actual_model_name = model
        
        print(f"[Faster-Whisper] 模型名称: {model}")
        print(f"[Faster-Whisper] 实际模型: {actual_model_name}")
        sys.stdout.flush()
        
        # 获取 whisper-faster 可执行文件路径（尝试多个位置）
        current_dir = Path(__file__).parent.parent  # app 目录
        
        # 尝试多个可能的路径
        possible_exe_paths = [
            current_dir / 'tools' / 'whisper-faster.exe',           # app/tools/
            current_dir / 'common' / 'tools' / 'whisper-faster.exe',  # app/common/tools/
            Path('whisper-faster.exe'),                             # 当前目录
        ]
        
        whisper_exe = None
        for path in possible_exe_paths:
            if path.exists():
                whisper_exe = path
                break
        
        # 检查可执行文件是否存在
        if not whisper_exe:
            error_msg = f"未找到 whisper-faster.exe，已尝试路径: {[str(p) for p in possible_exe_paths]}"
            print(f"[错误] {error_msg}")
            sys.stdout.flush()
            raise FileNotFoundError(error_msg)
        
        # 模型目录（尝试多个位置）
        possible_model_dirs = [
            current_dir / 'common' / 'models' / 'whisper-faster',
            current_dir / 'models' / 'whisper-faster',
        ]
        
        model_dir = None
        for path in possible_model_dirs:
            if path.exists():
                model_dir = path
                break
        
        if not model_dir:
            model_dir = possible_model_dirs[0]  # 使用第一个作为默认值
            print(f"[警告] 模型目录不存在，将使用: {model_dir}")
        
        # 构建相对路径（从当前工作目录到 exe）
        # 这样 PyInstaller 程序可以正确解析自己的资源
        try:
            relative_exe = whisper_exe.relative_to(Path.cwd())
            exe_path = str(relative_exe)
        except ValueError:
            # 如果无法计算相对路径，使用绝对路径
            exe_path = str(whisper_exe.absolute())
        
        print(f"[Faster-Whisper] ✓ 可执行文件: {exe_path}")
        print(f"[Faster-Whisper] 模型目录: {model_dir}")
        print(f"[Faster-Whisper] 将自动输出详细的 VAD 日志")
        sys.stdout.flush()
        
        # 检查本地是否有模型文件（仅用于提示）
        model_path = model_dir / f"faster-whisper-{actual_model_name}"
        if not model_path.exists():
            # 尝试不带前缀的路径
            model_path = model_dir / actual_model_name
        
        # ⚠️ 重要：whisper-faster.exe 需要的是模型名称，不是完整路径
        # 它会自动在工作目录（cwd）下查找模型或从 Hugging Face 下载
        model_arg = actual_model_name
        
        if model_path.exists() and model_path.is_dir():
            print(f"[Faster-Whisper] ✓ 本地模型目录: {model_path}")
            print(f"[Faster-Whisper] 使用模型名称: {model_arg}")
        else:
            print(f"[Faster-Whisper] ⚠ 本地未找到模型，将从 Hugging Face 下载: {model_arg}")
        
        # 构建命令参数（优化内存使用）
        cmd_args = [
            exe_path,
            '--beep_off',
            '--verbose', 'True',
            '--model', model_arg,  
            '--model_dir', str(model_dir.absolute()),  
            '--task', 'transcribe',
            '--language', language,
            '--output_format', 'srt',
            '--output_dir', str(wav_file.parent.absolute()),
            str(wav_file.absolute()),
            '--compute_type', 'float16'
        ]
        
        # 获取额外参数
        params = whisper_params.get('faster_whisper_params', '')
        if params:
            cmd_args.extend(params.split())
        
        print(f"[Faster-Whisper] 完整命令: {' '.join(cmd_args)}")
        sys.stdout.flush()
        
        return cmd_args
    
    def _generate_output(self, srt_file: Path, input_file: Path, 
                        output_format: str, kwargs: Dict[str, Any]) -> Path:
        """
        生成指定格式的输出文件
        
        Args:
            srt_file: SRT 文件路径
            input_file: 输入文件路径
            output_format: 输出格式
            kwargs: 配置参数字典
            
        Returns:
            输出文件路径
        """
        # 获取时间戳设置（默认为 False）
        include_timestamp = kwargs.get('include_timestamp', False)
        
        # 获取均分人数设置（默认为 0，表示不均分）
        split_parts = kwargs.get('split_parts', 0)
        
        # 获取保存文件夹（默认为输入文件所在目录）
        save_folder = kwargs.get('save_folder')
        if save_folder:
            save_folder = Path(save_folder)
            # 确保保存文件夹存在
            save_folder.mkdir(parents=True, exist_ok=True)
        else:
            save_folder = input_file.parent
        
        # 原文格式处理
        if output_format == OutputFormat.SRT_ORIGINAL:
            # 原文 SRT（始终包含时间戳，这是 SRT 格式的必需部分）
            output_file = save_folder / f"{input_file.stem}.srt"
            if srt_file != output_file:
                import shutil
                shutil.copy(srt_file, output_file)
            return output_file
        
        elif output_format == OutputFormat.LRC_ORIGINAL:
            # 原文 LRC（始终包含时间戳，这是 LRC 格式的必需部分）
            output_file = save_folder / f"{input_file.stem}.lrc"
            self._srt_to_lrc(srt_file, output_file)
            return output_file
        
        elif output_format == OutputFormat.TXT_ORIGINAL:
            # 原文 TXT
            output_file = save_folder / f"{input_file.stem}.txt"
            self._srt_to_txt(srt_file, output_file, include_timestamp=include_timestamp, split_parts=split_parts)
            return output_file
        
        elif output_format == OutputFormat.XLSX_ORIGINAL:
            # 原文 XLSX
            output_file = save_folder / f"{input_file.stem}.xlsx"
            self._srt_to_xlsx(srt_file, output_file, include_timestamp=include_timestamp, split_parts=split_parts)
            return output_file
        
        # 双语格式处理（需要翻译文件）
        elif output_format == OutputFormat.SRT_BILINGUAL:
            # 双语 SRT（始终包含时间戳）
            translated_srt = kwargs.get('translated_srt')
            if not translated_srt or not Path(translated_srt).exists():
                print("[WARNING] 未找到翻译文件，仅输出原文")
                return self._generate_output(srt_file, input_file, OutputFormat.SRT_ORIGINAL, kwargs)
            
            output_file = save_folder / f"{input_file.stem}_bilingual.srt"
            self._merge_bilingual_srt(srt_file, Path(translated_srt), output_file)
            return output_file
        
        elif output_format == OutputFormat.TXT_BILINGUAL:
            # 双语 TXT
            translated_srt = kwargs.get('translated_srt')
            if not translated_srt or not Path(translated_srt).exists():
                print("[WARNING] 未找到翻译文件，仅输出原文")
                return self._generate_output(srt_file, input_file, OutputFormat.TXT_ORIGINAL, kwargs)
            
            output_file = save_folder / f"{input_file.stem}_bilingual.txt"
            self._merge_bilingual_txt(srt_file, Path(translated_srt), output_file, 
                                      include_timestamp=include_timestamp, split_parts=split_parts)
            return output_file
        
        elif output_format == OutputFormat.XLSX_BILINGUAL:
            # 双语 XLSX
            translated_srt = kwargs.get('translated_srt')
            if not translated_srt or not Path(translated_srt).exists():
                print("[WARNING] 未找到翻译文件，仅输出原文")
                return self._generate_output(srt_file, input_file, OutputFormat.XLSX_ORIGINAL, kwargs)
            
            output_file = save_folder / f"{input_file.stem}_bilingual.xlsx"
            self._merge_bilingual_xlsx(srt_file, Path(translated_srt), output_file,
                                       include_timestamp=include_timestamp, split_parts=split_parts)
            return output_file
        
        else:
            # 默认返回 SRT
            print(f"[WARNING] 未知的输出格式: {output_format}，使用默认 SRT")
            return srt_file
    
    def _convert_srt_format(self, srt_file: Path, output_format: str, 
                           kwargs: Dict[str, Any]) -> Dict[str, str]:
        """
        转换 SRT 文件格式（用于直接输入 SRT 文件的情况）
        
        Args:
            srt_file: SRT 文件路径
            output_format: 输出格式
            kwargs: 配置参数字典
            
        Returns:
            结果字典
        """
        output_path = self._generate_output(srt_file, srt_file, output_format, kwargs)
        
        return {
            'output_path': str(output_path),
            'srt_path': str(srt_file)
        }
    
    # ==================== 工具函数 ====================
    
    @staticmethod
    def _calculate_time_splits(total_duration: float, split_parts: int) -> list:
        """
        计算时间段划分
        
        算法说明：
        1. 将总时长向上取整为分钟数
        2. 判断取整后的分钟数能否被人数整除
        3. 如果能整除，按整除结果均分
        4. 如果不能整除，前面的part取整数分钟，余数给最后一人
        
        Args:
            total_duration: 总时长（秒）
            split_parts: 划分份数
            
        Returns:
            时间段列表，每个元素为 (start, end) 元组（秒）
        """
        if split_parts <= 0:
            return [(0, total_duration)]
        
        # 转换为分钟
        total_minutes = total_duration / 60.0
        
        # 向上取整（如果有小数部分）
        import math
        total_minutes_ceil = math.ceil(total_minutes)
        
        # 如果总时长小于1分钟，按1分钟算
        if total_minutes_ceil < 1:
            total_minutes_ceil = 1
        
        # 判断能否整除
        if total_minutes_ceil % split_parts == 0:
            # 能整除：均分
            minutes_per_part = total_minutes_ceil // split_parts
            
            print(f"[时间划分] 总时长 {total_minutes:.2f} 分钟向上取整为 {total_minutes_ceil} 分钟")
            print(f"[时间划分] {total_minutes_ceil} 分钟可被 {split_parts} 人整除，每人 {minutes_per_part} 分钟")
            
            time_splits = []
            for i in range(split_parts):
                start = i * minutes_per_part * 60.0
                if i == split_parts - 1:
                    # 最后一段：到实际结束时间
                    end = total_duration
                else:
                    end = (i + 1) * minutes_per_part * 60.0
                
                time_splits.append((start, end))
        else:
            # 不能整除：前面取整，余数给最后一人
            minutes_per_part = total_minutes_ceil // split_parts
            remainder = total_minutes_ceil % split_parts
            
            print(f"[时间划分] 总时长 {total_minutes:.2f} 分钟向上取整为 {total_minutes_ceil} 分钟")
            print(f"[时间划分] {total_minutes_ceil} 分钟不能被 {split_parts} 人整除")
            print(f"[时间划分] 前 {split_parts-1} 人每人 {minutes_per_part} 分钟，最后一人分配剩余时间")
            
            time_splits = []
            for i in range(split_parts):
                start = i * minutes_per_part * 60.0
                if i == split_parts - 1:
                    # 最后一段：到实际结束时间（包含余数）
                    end = total_duration
                else:
                    end = (i + 1) * minutes_per_part * 60.0
                
                time_splits.append((start, end))
        
        return time_splits
    
    @staticmethod
    def _format_timestamp_srt(seconds: float) -> str:
        """格式化为SRT时间戳格式 HH:MM:SS,mmm"""
        total_seconds = int(seconds)
        hh, remainder = divmod(total_seconds, 3600)
        mm, ss = divmod(remainder, 60)
        ms = int((seconds * 1000) % 1000)
        return f"{hh:02d}:{mm:02d}:{ss:02d},{ms:03d}"
    
    @staticmethod
    def _format_timestamp_lrc(seconds: float) -> str:
        """格式化为LRC时间戳格式 MM:SS.mmm"""
        total_seconds = int(seconds)
        mm, ss = divmod(total_seconds, 60)
        ms = int((seconds * 1000) % 1000)
        return f"{mm:02d}:{ss:02d}.{ms:03d}"
    
    @staticmethod
    def _parse_srt_timestamp(timestamp: str) -> float:
        """解析SRT时间戳为秒数"""
        # 格式: HH:MM:SS,mmm
        time_parts = timestamp.replace(',', ':').split(':')
        if len(time_parts) == 4:
            hh, mm, ss, ms = map(int, time_parts)
            return hh * 3600 + mm * 60 + ss + ms / 1000.0
        return 0.0
    
    @staticmethod
    def _get_unique_filename(filepath: Path) -> Path:
        """生成唯一的文件名，如果文件已存在则在文件名后添加数字后缀"""
        if not filepath.exists():
            return filepath
        
        directory = filepath.parent
        name = filepath.stem
        ext = filepath.suffix
        
        counter = 1
        while True:
            new_filepath = directory / f"{name}_{counter}{ext}"
            if not new_filepath.exists():
                return new_filepath
            counter += 1
    
    @staticmethod
    def replace_punctuation_with_space(text: str) -> str:
        """
        将文本中的标点符号替换为空格
        支持中文、英文、日文等多种标点符号
        
        Args:
            text: 输入文本
            
        Returns:
            处理后的文本
        """
        if not text:
            return text
        
        # 定义要替换的标点符号（包括中文、英文、日文等）
        punctuation_pattern = r'[。！？，、；：""''（）【】〈〉〔〕｛｝,.!?;:\'"()[\]{}<>/@#$%^&*+=|\\~`_-]'
        
        # 将标点符号替换为空格
        result = re.sub(punctuation_pattern, '　', text)
        
        # 将多个连续空格替换为单个空格，并去除首尾空格
        result = re.sub(r'\s+', ' ', result).strip()
        
        return result
    
    def _parse_srt(self, srt_file: Path) -> list:
        """
        解析 SRT 文件
        
        Returns:
            包含字幕段的列表，每个段包含:
            - start: 开始时间（秒，float）
            - end: 结束时间（秒，float）
            - start_str: 开始时间字符串（SRT格式）
            - end_str: 结束时间字符串（SRT格式）
            - text: 字幕文本
        """
        segments = []
        
        with open(srt_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 简单的 SRT 解析
        blocks = content.strip().split('\n\n')
        
        for block in blocks:
            lines = block.strip().split('\n')
            if len(lines) >= 3:
                # 时间戳行
                time_line = lines[1]
                match = re.match(r'(\d{2}:\d{2}:\d{2},\d{3}) --> (\d{2}:\d{2}:\d{2},\d{3})', time_line)
                if match:
                    start_str, end_str = match.groups()
                    text = '\n'.join(lines[2:])
                    segments.append({
                        'start': self._parse_srt_timestamp(start_str),
                        'end': self._parse_srt_timestamp(end_str),
                        'start_str': start_str,
                        'end_str': end_str,
                        'text': text
                    })
        
        return segments
    
    # ==================== 格式转换方法 ====================
    
    def _srt_to_lrc(self, srt_file: Path, lrc_file: Path):
        """将 SRT 文件转换为 LRC 格式"""
        segments = self._parse_srt(srt_file)
        
        print(f"[转换] 正在生成 LRC 文件: {lrc_file.name}")
        
        with open(lrc_file, 'w', encoding='utf-8') as f:
            for seg in segments:
                # 替换标点符号为空格
                text = self.replace_punctuation_with_space(seg['text'])
                # 使用工具函数格式化时间戳
                timestamp = self._format_timestamp_lrc(seg['start'])
                f.write(f"[{timestamp}] {text}\n")
        
        print(f"[转换] LRC 文件生成完成")

    def _srt_to_txt(self, srt_file: Path, txt_file: Path, include_timestamp: bool = True, split_parts: int = 0):
        """将 SRT 文件转换为 TXT 格式"""
        segments = self._parse_srt(srt_file)
        
        # 如果不需要均分，直接生成一个文件
        if split_parts <= 0:
            print(f"[转换] 正在生成 TXT 文件: {txt_file.name}")
            
            with open(txt_file, 'w', encoding='utf-8') as f:
                for seg in segments:
                    # 替换标点符号为空格
                    text = self.replace_punctuation_with_space(seg['text'])
                    
                    if include_timestamp:
                        f.write(f"[{seg['start_str']} --> {seg['end_str']}]\n")
                    f.write(f"{text}\n\n")
            
            print(f"[转换] TXT 文件生成完成")
            return
        
        # 需要均分：计算时间段并生成多个文件
        if not segments:
            print(f"[警告] 没有字幕段，跳过均分")
            return
        
        # 获取总时长
        total_duration = segments[-1]['end']
        
        # 计算时间段划分
        time_splits = self._calculate_time_splits(total_duration, split_parts)
        
        print(f"[转换] 正在按 {split_parts} 人均分生成 TXT 文件...")
        print(f"[转换] 总时长: {total_duration:.2f} 秒 ({total_duration/60:.2f} 分钟)")
        
        # 为每个时间段生成文件
        for part_idx, (start_time, end_time) in enumerate(time_splits, 1):
            # 生成文件名：原文件名_part_1.txt
            part_filename = f"{txt_file.stem}_part_{part_idx}{txt_file.suffix}"
            part_file = txt_file.with_name(part_filename)
            
            # 过滤属于该时间段的segments
            part_segments = [
                seg for seg in segments 
                if seg['start'] >= start_time and seg['start'] < end_time
            ]
            
            print(f"[转换] Part {part_idx}: {start_time/60:.2f}~{end_time/60:.2f} 分钟, {len(part_segments)} 条字幕 -> {part_file.name}")
            
            # 写入文件
            with open(part_file, 'w', encoding='utf-8') as f:
                # 写入时间段信息
                f.write(f"# Part {part_idx} of {split_parts}\n")
                f.write(f"# Time Range: {self._format_timestamp_srt(start_time)} --> {self._format_timestamp_srt(end_time)}\n\n")
                
                for seg in part_segments:
                    # 替换标点符号为空格
                    text = self.replace_punctuation_with_space(seg['text'])
                    
                    if include_timestamp:
                        f.write(f"[{seg['start_str']} --> {seg['end_str']}]\n")
                    f.write(f"{text}\n\n")
        
        print(f"[转换] TXT 文件均分完成，共生成 {split_parts} 个文件")

    def _srt_to_xlsx(self, srt_file: Path, xlsx_file: Path, include_timestamp: bool = True, split_parts: int = 0):
        """将 SRT 文件转换为 XLSX 格式"""
        segments = self._parse_srt(srt_file)
        
        # 如果不需要均分，直接生成一个文件
        if split_parts <= 0:
            print(f"[转换] 正在生成 XLSX 文件: {xlsx_file.name}")
            
            # 避免文件重名
            unique_xlsx_file = self._get_unique_filename(xlsx_file)
            if unique_xlsx_file != xlsx_file:
                print(f"[INFO] 文件重名，自动重命名为: {unique_xlsx_file.name}")
            
            try:
                # 创建工作簿
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "字幕"
                
                # 根据时间戳设置设置表头
                if include_timestamp:
                    headers = ['序号', '开始时间', '结束时间', '字幕内容', '翻译', '屏幕字']
                else:
                    headers = ['序号', '字幕内容', '翻译', '屏幕字']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # 写入数据
                for idx, seg in enumerate(segments, 1):
                    # 替换标点符号为空格
                    text = self.replace_punctuation_with_space(seg['text'])
                    
                    if include_timestamp:
                        ws.cell(row=idx+1, column=1, value=idx)
                        ws.cell(row=idx+1, column=2, value=seg['start_str'])
                        ws.cell(row=idx+1, column=3, value=seg['end_str'])
                        ws.cell(row=idx+1, column=4, value=text)
                    else:
                        ws.cell(row=idx+1, column=1, value=idx)
                        ws.cell(row=idx+1, column=2, value=text)
                
                # 调整列宽
                if include_timestamp:
                    ws.column_dimensions['A'].width = 8
                    ws.column_dimensions['B'].width = 15
                    ws.column_dimensions['C'].width = 15
                    ws.column_dimensions['D'].width = 60
                else:
                    ws.column_dimensions['A'].width = 8
                    ws.column_dimensions['B'].width = 60
                
                # 保存
                wb.save(unique_xlsx_file)
                print(f"[转换] XLSX 文件生成成功: {unique_xlsx_file.name}")
                
            except PermissionError as e:
                error_msg = f"文件权限错误，无法写入 XLSX 文件: {unique_xlsx_file}"
                print(f"[ERROR] {error_msg}")
                raise RuntimeError(error_msg)
            except Exception as e:
                error_msg = f"生成 XLSX 文件时发生错误: {str(e)}"
                print(f"[ERROR] {error_msg}")
                raise RuntimeError(error_msg)
            return
        
        # 需要均分：计算时间段并生成多个文件
        if not segments:
            print(f"[警告] 没有字幕段，跳过均分")
            return
        
        # 获取总时长
        total_duration = segments[-1]['end']
        
        # 计算时间段划分
        time_splits = self._calculate_time_splits(total_duration, split_parts)
        
        print(f"[转换] 正在按 {split_parts} 人均分生成 XLSX 文件...")
        print(f"[转换] 总时长: {total_duration:.2f} 秒 ({total_duration/60:.2f} 分钟)")
        
        # 为每个时间段生成文件
        for part_idx, (start_time, end_time) in enumerate(time_splits, 1):
            # 生成文件名：原文件名_part_1.xlsx
            part_filename = f"{xlsx_file.stem}_part_{part_idx}{xlsx_file.suffix}"
            part_file = xlsx_file.with_name(part_filename)
            
            # 避免文件重名
            unique_part_file = self._get_unique_filename(part_file)
            
            # 过滤属于该时间段的segments
            part_segments = [
                seg for seg in segments 
                if seg['start'] >= start_time and seg['start'] < end_time
            ]
            
            print(f"[转换] Part {part_idx}: {start_time/60:.2f}~{end_time/60:.2f} 分钟, {len(part_segments)} 条字幕 -> {unique_part_file.name}")
            
            try:
                # 创建工作簿
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"字幕_Part{part_idx}"
                
                # 根据时间戳设置设置表头
                if include_timestamp:
                    headers = ['序号', '开始时间', '结束时间', '字幕内容', '翻译', '屏幕字']
                else:
                    headers = ['序号', '字幕内容', '翻译', '屏幕字']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # 写入数据
                for idx, seg in enumerate(part_segments, 1):
                    # 替换标点符号为空格
                    text = self.replace_punctuation_with_space(seg['text'])
                    
                    if include_timestamp:
                        ws.cell(row=idx+1, column=1, value=idx)
                        ws.cell(row=idx+1, column=2, value=seg['start_str'])
                        ws.cell(row=idx+1, column=3, value=seg['end_str'])
                        ws.cell(row=idx+1, column=4, value=text)
                    else:
                        ws.cell(row=idx+1, column=1, value=idx)
                        ws.cell(row=idx+1, column=2, value=text)
                
                # 调整列宽
                if include_timestamp:
                    ws.column_dimensions['A'].width = 8
                    ws.column_dimensions['B'].width = 15
                    ws.column_dimensions['C'].width = 15
                    ws.column_dimensions['D'].width = 60
                else:
                    ws.column_dimensions['A'].width = 8
                    ws.column_dimensions['B'].width = 60
                
                # 保存
                wb.save(unique_part_file)
                
            except PermissionError as e:
                error_msg = f"文件权限错误，无法写入 XLSX 文件: {unique_part_file}"
                print(f"[ERROR] {error_msg}")
                raise RuntimeError(error_msg)
            except Exception as e:
                error_msg = f"生成 XLSX 文件时发生错误: {str(e)}"
                print(f"[ERROR] {error_msg}")
                raise RuntimeError(error_msg)
        
        print(f"[转换] XLSX 文件均分完成，共生成 {split_parts} 个文件")

    def _merge_bilingual_srt(self, original_srt: Path, translated_srt: Path, output_srt: Path):
        """合并原文和翻译生成双语 SRT"""
        original_segments = self._parse_srt(original_srt)
        translated_segments = self._parse_srt(translated_srt)
        
        print(f"[转换] 正在生成双语 SRT 文件: {output_srt.name}")
        
        with open(output_srt, 'w', encoding='utf-8') as f:
            for idx, (orig, trans) in enumerate(zip(original_segments, translated_segments), 1):
                f.write(f"{idx}\n")
                f.write(f"{orig['start_str']} --> {orig['end_str']}\n")
                f.write(f"{orig['text']}\n")
                f.write(f"{trans['text']}\n")
                f.write("\n")
        
        print(f"[转换] 双语 SRT 文件生成完成")

    def _merge_bilingual_txt(self, original_srt: Path, translated_srt: Path, 
                             output_txt: Path, include_timestamp: bool = True, split_parts: int = 0):
        """合并原文和翻译生成双语 TXT"""
        original_segments = self._parse_srt(original_srt)
        translated_segments = self._parse_srt(translated_srt)
        
        # 如果不需要均分，直接生成一个文件
        if split_parts <= 0:
            print(f"[转换] 正在生成双语 TXT 文件: {output_txt.name}")
            
            with open(output_txt, 'w', encoding='utf-8') as f:
                for orig, trans in zip(original_segments, translated_segments):
                    # 替换标点符号为空格
                    orig_text = self.replace_punctuation_with_space(orig['text'])
                    trans_text = self.replace_punctuation_with_space(trans['text'])
                    
                    if include_timestamp:
                        f.write(f"[{orig['start_str']} --> {orig['end_str']}]\n")
                    f.write(f"原文: {orig_text}\n")
                    f.write(f"译文: {trans_text}\n\n")
            
            print(f"[转换] 双语 TXT 文件生成完成")
            return
        
        # 需要均分：计算时间段并生成多个文件
        if not original_segments:
            print(f"[警告] 没有字幕段，跳过均分")
            return
        
        # 获取总时长
        total_duration = original_segments[-1]['end']
        
        # 计算时间段划分
        time_splits = self._calculate_time_splits(total_duration, split_parts)
        
        print(f"[转换] 正在按 {split_parts} 人均分生成双语 TXT 文件...")
        print(f"[转换] 总时长: {total_duration:.2f} 秒 ({total_duration/60:.2f} 分钟)")
        
        # 为每个时间段生成文件
        for part_idx, (start_time, end_time) in enumerate(time_splits, 1):
            # 生成文件名：原文件名_part_1.txt
            part_filename = f"{output_txt.stem}_part_{part_idx}{output_txt.suffix}"
            part_file = output_txt.with_name(part_filename)
            
            # 过滤属于该时间段的segments
            part_pairs = [
                (orig, trans) for orig, trans in zip(original_segments, translated_segments)
                if orig['start'] >= start_time and orig['start'] < end_time
            ]
            
            print(f"[转换] Part {part_idx}: {start_time/60:.2f}~{end_time/60:.2f} 分钟, {len(part_pairs)} 条字幕 -> {part_file.name}")
            
            # 写入文件
            with open(part_file, 'w', encoding='utf-8') as f:
                # 写入时间段信息
                f.write(f"# Part {part_idx} of {split_parts}\n")
                f.write(f"# Time Range: {self._format_timestamp_srt(start_time)} --> {self._format_timestamp_srt(end_time)}\n\n")
                
                for orig, trans in part_pairs:
                    # 替换标点符号为空格
                    orig_text = self.replace_punctuation_with_space(orig['text'])
                    trans_text = self.replace_punctuation_with_space(trans['text'])
                    
                    if include_timestamp:
                        f.write(f"[{orig['start_str']} --> {orig['end_str']}]\n")
                    f.write(f"原文: {orig_text}\n")
                    f.write(f"译文: {trans_text}\n\n")
        
        print(f"[转换] 双语 TXT 文件均分完成，共生成 {split_parts} 个文件")

    def _merge_bilingual_xlsx(self, original_srt: Path, translated_srt: Path, 
                              output_xlsx: Path, include_timestamp: bool = True, split_parts: int = 0):
        """合并原文和翻译生成双语 XLSX"""
        original_segments = self._parse_srt(original_srt)
        translated_segments = self._parse_srt(translated_srt)
        
        # 如果不需要均分，直接生成一个文件
        if split_parts <= 0:
            print(f"[转换] 正在生成双语 XLSX 文件: {output_xlsx.name}")
            
            # 避免文件重名
            unique_xlsx_file = self._get_unique_filename(output_xlsx)
            if unique_xlsx_file != output_xlsx:
                print(f"[INFO] 文件重名，自动重命名为: {unique_xlsx_file.name}")
            
            try:
                # 创建工作簿
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "双语字幕"
                
                # 根据时间戳设置设置表头
                if include_timestamp:
                    headers = ['序号', '开始时间', '结束时间', 'Original（原文）', 'Translate（译文）']
                else:
                    headers = ['序号', 'Original（原文）', 'Translate（译文）']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # 写入数据
                for idx, (orig, trans) in enumerate(zip(original_segments, translated_segments), 1):
                    # 替换标点符号为空格
                    orig_text = self.replace_punctuation_with_space(orig['text'])
                    trans_text = self.replace_punctuation_with_space(trans['text'])
                    
                    if include_timestamp:
                        ws.cell(row=idx+1, column=1, value=idx)
                        ws.cell(row=idx+1, column=2, value=orig['start_str'])
                        ws.cell(row=idx+1, column=3, value=orig['end_str'])
                        ws.cell(row=idx+1, column=4, value=orig_text)
                        ws.cell(row=idx+1, column=5, value=trans_text)
                    else:
                        ws.cell(row=idx+1, column=1, value=idx)
                        ws.cell(row=idx+1, column=2, value=orig_text)
                        ws.cell(row=idx+1, column=3, value=trans_text)
                
                # 调整列宽
                if include_timestamp:
                    ws.column_dimensions['A'].width = 8
                    ws.column_dimensions['B'].width = 15
                    ws.column_dimensions['C'].width = 15
                    ws.column_dimensions['D'].width = 40
                    ws.column_dimensions['E'].width = 40
                else:
                    ws.column_dimensions['A'].width = 8
                    ws.column_dimensions['B'].width = 40
                    ws.column_dimensions['C'].width = 40
                
                # 保存
                wb.save(unique_xlsx_file)
                print(f"[转换] 双语 XLSX 文件生成成功: {unique_xlsx_file.name}")
                
            except PermissionError as e:
                error_msg = f"文件权限错误，无法写入 XLSX 文件: {unique_xlsx_file}"
                print(f"[ERROR] {error_msg}")
                raise RuntimeError(error_msg)
            except Exception as e:
                error_msg = f"生成双语 XLSX 文件时发生错误: {str(e)}"
                print(f"[ERROR] {error_msg}")
                raise RuntimeError(error_msg)
            return
        
        # 需要均分：计算时间段并生成多个文件
        if not original_segments:
            print(f"[警告] 没有字幕段，跳过均分")
            return
        
        # 获取总时长
        total_duration = original_segments[-1]['end']
        
        # 计算时间段划分
        time_splits = self._calculate_time_splits(total_duration, split_parts)
        
        print(f"[转换] 正在按 {split_parts} 人均分生成双语 XLSX 文件...")
        print(f"[转换] 总时长: {total_duration:.2f} 秒 ({total_duration/60:.2f} 分钟)")
        
        # 为每个时间段生成文件
        for part_idx, (start_time, end_time) in enumerate(time_splits, 1):
            # 生成文件名：原文件名_part_1.xlsx
            part_filename = f"{output_xlsx.stem}_part_{part_idx}{output_xlsx.suffix}"
            part_file = output_xlsx.with_name(part_filename)
            
            # 避免文件重名
            unique_part_file = self._get_unique_filename(part_file)
            
            # 过滤属于该时间段的segments
            part_pairs = [
                (orig, trans) for orig, trans in zip(original_segments, translated_segments)
                if orig['start'] >= start_time and orig['start'] < end_time
            ]
            
            print(f"[转换] Part {part_idx}: {start_time/60:.2f}~{end_time/60:.2f} 分钟, {len(part_pairs)} 条字幕 -> {unique_part_file.name}")
            
            try:
                # 创建工作簿
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"双语字幕_Part{part_idx}"
                
                # 根据时间戳设置设置表头
                if include_timestamp:
                    headers = ['序号', '开始时间', '结束时间', 'Original（原文）', 'Translate（译文）']
                else:
                    headers = ['序号', 'Original（原文）', 'Translate（译文）']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # 写入数据
                for idx, (orig, trans) in enumerate(part_pairs, 1):
                    # 替换标点符号为空格
                    orig_text = self.replace_punctuation_with_space(orig['text'])
                    trans_text = self.replace_punctuation_with_space(trans['text'])
                    
                    if include_timestamp:
                        ws.cell(row=idx+1, column=1, value=idx)
                        ws.cell(row=idx+1, column=2, value=orig['start_str'])
                        ws.cell(row=idx+1, column=3, value=orig['end_str'])
                        ws.cell(row=idx+1, column=4, value=orig_text)
                        ws.cell(row=idx+1, column=5, value=trans_text)
                    else:
                        ws.cell(row=idx+1, column=1, value=idx)
                        ws.cell(row=idx+1, column=2, value=orig_text)
                        ws.cell(row=idx+1, column=3, value=trans_text)
                
                # 调整列宽
                if include_timestamp:
                    ws.column_dimensions['A'].width = 8
                    ws.column_dimensions['B'].width = 15
                    ws.column_dimensions['C'].width = 15
                    ws.column_dimensions['D'].width = 40
                    ws.column_dimensions['E'].width = 40
                else:
                    ws.column_dimensions['A'].width = 8
                    ws.column_dimensions['B'].width = 40
                    ws.column_dimensions['C'].width = 40
                
                # 保存
                wb.save(unique_part_file)
                
            except PermissionError as e:
                error_msg = f"文件权限错误，无法写入 XLSX 文件: {unique_part_file}"
                print(f"[ERROR] {error_msg}")
                raise RuntimeError(error_msg)
            except Exception as e:
                error_msg = f"生成双语 XLSX 文件时发生错误: {str(e)}"
                print(f"[ERROR] {error_msg}")
                raise RuntimeError(error_msg)
        
        print(f"[转换] 双语 XLSX 文件均分完成，共生成 {split_parts} 个文件")


# 全局服务实例
transcriptionService = TranscriptionService()